#coding=utf-8
from openpyxl import load_workbook
from openpyxl import Workbook
#d = ws.cell(row = 3, column = 0)
workbook_ = load_workbook('e.xlsx') #导入工作表
sheetnames =workbook_.get_sheet_names() #获得表单名字
sheet = workbook_.get_sheet_by_name(sheetnames[0])   #从工作表中提取某一表单
sheet = workbook_.active    #新加入的一行
arrdata=[]  #定义了一个列表
sum = 0
socelist = [1.5,2,2,0.5,3,3,4.5,1,4.5,5,2,1,2,2,4,3,4.5,4,1,3,1.5,2]
i = 0
for rowNum in range(3,31):
    for colNum in range(4,26):
        data = sheet.cell(row=rowNum,column=colNum).value
        arrdata.append(data)
    for item in arrdata:
        item = item * socelist[i]
        sum += item
        i = i+1
    print(sum)
    sum = sum / 57
    print(sum)
    #sheet['Z4'] = sum
    sheet.cell(row=rowNum, column=26, value='%f' % sum)
    #for cell in sheet['Z4':'Z30']:
     #   cell = sum
    sum=0
    i = 0
    arrdata = []
workbook_.save(filename="e.xlsx")






